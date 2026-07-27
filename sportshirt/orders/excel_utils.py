"""
ยูทิลิตี้สำหรับซิงค์ข้อมูลคำสั่งซื้อ (Order) ลงไฟล์ Excel (.xlsx)

แนวคิด:
- Django (ฐานข้อมูลหลัก / source of truth) จะบันทึกคำสั่งซื้อทุกครั้งที่มีลูกค้ากรอกฟอร์ม
- ทุกครั้งที่บันทึกสำเร็จ ระบบจะ "append" ข้อมูลแถวใหม่ลงไฟล์ Excel ที่ path คงที่
  (ORDERS_EXCEL_PATH ใน settings.py) โดยอัตโนมัติ
- ไฟล์ Excel นี้คือไฟล์ที่ UiPath Studio จะเปิดอ่าน/monitor เพื่อทำงานอัตโนมัติต่อ
  (เช่น "Read Range" หรือ "Excel Application Scope" -> ประมวลผลออเดอร์ -> พิมพ์ใบปะหน้า ฯลฯ)
- คอลัมน์ของไฟล์ตรงกับที่ผู้ใช้ระบุไว้:
    ชื่อ-นามสกุล | เบอร์โทรศัพท์ | จำนวนรวม (ตัว) | ราคารวม (บาท) |
    วันที่ต้องการรับสินค้า | วันที่สั่งซื้อ
- ฟังก์ชันนี้ thread-safe ในระดับพื้นฐานด้วย file lock ง่าย ๆ (filelock) เพื่อกันข้อมูลชนกัน
  เวลามีคำสั่งซื้อเข้ามาพร้อมกันหลายรายการ
"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.conf import settings

EXCEL_HEADERS = [
    "ชื่อ-นามสกุล",
    "อีเมล",
    "ไซส์เสื้อ",
    "เบอร์โทรศัพท์",
    "จำนวนรวม (ตัว)",
    "ราคารวม (บาท)",
    "ช่องทางชำระเงิน",
    "สถานะชำระเงิน",
    "วันที่ต้องการรับสินค้า",
    "วันที่สั่งซื้อ",
    "ที่อยู่จัดส่ง",
    "Order ID",
]

# ความกว้างคอลัมน์ (จำนวนตัวอักษร) ให้เหมาะกับเนื้อหาแต่ละคอลัมน์
COLUMN_WIDTHS = [22, 26, 10, 15, 14, 14, 18, 16, 20, 20, 45, 10]

_HEADER_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)


def _style_worksheet(ws):
    """จัดสไตล์ชีต: หัวตารางพื้นเหลืองตัวหนา + ความกว้างคอลัมน์ที่เหมาะสม + เส้นขอบบาง ๆ"""
    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 24

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def _get_excel_path():
    path = settings.ORDERS_EXCEL_PATH
    os.makedirs(os.path.dirname(path), exist_ok=True)
    return path


def _ensure_workbook(path):
    """สร้างไฟล์ใหม่พร้อม header แถวแรก ถ้ายังไม่มีไฟล์ หรือไฟล์เสีย"""
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
            return wb
        except Exception:
            # ไฟล์เสีย/เปิดไม่ได้ -> สร้างใหม่ (สำรองไฟล์เดิมไว้กันข้อมูลหาย)
            backup_path = path + ".corrupt.bak"
            os.replace(path, backup_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "ข้อมูล"
    ws.append(EXCEL_HEADERS)
    _style_worksheet(ws)
    return wb


def append_order_to_excel(order):
    """
    เพิ่มแถวใหม่ 1 แถวสำหรับ order ที่บันทึกสำเร็จ
    เรียกใช้ทันทีหลังจาก order.save() ใน view
    """
    path = _get_excel_path()

    # ใช้ file lock ง่าย ๆ ผ่าน filelock library ถ้ามี ไม่มีก็ข้ามไป (เหมาะกับงานปริมาณไม่สูงมาก)
    try:
        from filelock import FileLock
        lock = FileLock(path + ".lock", timeout=10)
    except ImportError:
        class _NoLock:
            def __enter__(self):
                return self

            def __exit__(self, *a):
                return False
        lock = _NoLock()

    with lock:
        wb = _ensure_workbook(path)
        ws = wb["ข้อมูล"] if "ข้อมูล" in wb.sheetnames else wb.active
        ws.append([
            order.full_name,
            order.email,
            order.shirt_size,
            order.phone,
            order.quantity,
            order.total_price,
            order.get_payment_method_display(),
            order.get_payment_status_display(),
            order.expected_delivery_date.strftime("%d/%m/%Y"),
            order.order_date.strftime("%d/%m/%Y %H:%M"),
            order.full_address,
            order.pk,
        ])
        _style_worksheet(ws)
        wb.save(path)

    order.synced_to_excel = True
    order.save(update_fields=["synced_to_excel"])
    return path


def resync_all_orders():
    """
    สร้างไฟล์ Excel ใหม่ทั้งหมดจากข้อมูลใน Django database
    มีประโยชน์เวลาไฟล์ Excel หาย/เสีย หรือถูก UiPath ลบแถวออกไปแล้วต้องการดึงใหม่ทั้งหมด
    เรียกผ่าน management command: python manage.py resync_excel
    """
    from .models import Order

    path = _get_excel_path()
    wb = Workbook()
    ws = wb.active
    ws.title = "ข้อมูล"
    ws.append(EXCEL_HEADERS)

    for order in Order.objects.order_by("order_date"):
        ws.append([
            order.full_name,
            order.email,
            order.shirt_size,
            order.phone,
            order.quantity,
            order.total_price,
            order.get_payment_method_display(),
            order.get_payment_status_display(),
            order.expected_delivery_date.strftime("%d/%m/%Y"),
            order.order_date.strftime("%d/%m/%Y %H:%M"),
            order.full_address,
            order.pk,
        ])
    _style_worksheet(ws)
    wb.save(path)
    Order.objects.update(synced_to_excel=True)
    return path
